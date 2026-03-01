#!/usr/bin/env python3
"""
既存 mrt_predictions.xlsx の見解（キャッチコピーのみ）を
DB内の見解全文（YosoKenkaiTitle + YosoKenkaiTxt）で上書きする。

使い方:
  python3 scripts/enrich_comments.py
  python3 scripts/enrich_comments.py --dry-run   # 変更プレビューのみ
  python3 scripts/enrich_comments.py --xlsx /path/to/file.xlsx
"""

import argparse
import json
import re
import sqlite3
from pathlib import Path

import openpyxl

# パス
DB_PATH = Path(__file__).resolve().parent.parent / "data" / "mrt_predictions.db"
EXISTING_XLSX = Path(
    "/mnt/c/Users/owner/Documents/Obsidian Vault/10_Projects/"
    "keirin_prediction/data/mrt_predictions.xlsx"
)


def normalize_venue(race_name: str) -> tuple:
    """レース名（例: 広島1R）から (会場, レース番号R) を抽出。"""
    m = re.match(r'([\u4e00-\u9fff\u3040-\u309f\u30a0-\u30ff]+?)(\d{1,2}R)', race_name)
    if m:
        return m.group(1), m.group(2)
    return race_name, ""


def normalize_date(date_str: str) -> str:
    """日付を YYYY-MM-DD 形式に統一。"""
    if not date_str:
        return ""
    # 2025/12/20 → 2025-12-20
    date_str = str(date_str).strip()
    m = re.match(r'(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})', date_str)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return date_str


def main():
    parser = argparse.ArgumentParser(description="既存xlsxに見解全文を補完")
    parser.add_argument("--xlsx", type=str, default=str(EXISTING_XLSX),
                        help="対象xlsxパス")
    parser.add_argument("--dry-run", action="store_true",
                        help="変更をプレビューのみ（保存しない）")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERROR: {xlsx_path} not found")
        return

    if not DB_PATH.exists():
        print(f"ERROR: {DB_PATH} not found")
        return

    # DB読み込み
    conn = sqlite3.connect(str(DB_PATH))
    rows = conn.execute("""
        SELECT date, venue, race_number, comment
        FROM mrt_predictions
        WHERE comment IS NOT NULL AND comment != ''
    """).fetchall()
    conn.close()

    # DBデータのインデックス: (date, venue, race_number) → comment
    db_comments = {}
    for date, venue, race_num, comment in rows:
        key = (date, venue, race_num)
        db_comments[key] = comment
    print(f"DB: {len(db_comments)} predictions with comments loaded")

    # xlsx読み込み
    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb.active

    # ヘッダー確認
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    print(f"xlsx headers: {headers}")

    # 見解は8列目（H列）
    comment_col = None
    for i, h in enumerate(headers, 1):
        if h and "見解" in str(h):
            comment_col = i
            break
    if not comment_col:
        print("ERROR: 見解 column not found")
        return
    print(f"見解 column: {comment_col}")

    # 日付=1, レース名=2
    date_col = 1
    race_col = 2

    enriched = 0
    skipped = 0
    not_found = 0

    for row_idx in range(2, ws.max_row + 1):
        date_val = ws.cell(row=row_idx, column=date_col).value
        race_val = ws.cell(row=row_idx, column=race_col).value
        old_comment = ws.cell(row=row_idx, column=comment_col).value or ""

        date_norm = normalize_date(str(date_val))
        venue, race_num = normalize_venue(str(race_val))

        key = (date_norm, venue, race_num)
        new_comment = db_comments.get(key)

        if new_comment and len(new_comment) > len(old_comment) + 10:
            if args.dry_run:
                print(f"  Row {row_idx}: {date_norm} {venue}{race_num}")
                print(f"    OLD ({len(old_comment)}): {old_comment[:60]}...")
                print(f"    NEW ({len(new_comment)}): {new_comment[:60]}...")
            else:
                ws.cell(row=row_idx, column=comment_col).value = new_comment
            enriched += 1
        elif new_comment:
            skipped += 1
        else:
            not_found += 1

    print(f"\nResults: {enriched} enriched, {skipped} already ok, "
          f"{not_found} not in DB")

    if not args.dry_run and enriched > 0:
        wb.save(str(xlsx_path))
        print(f"Saved to {xlsx_path}")
    elif args.dry_run:
        print("(dry-run: no changes saved)")


if __name__ == "__main__":
    main()
