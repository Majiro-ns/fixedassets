#!/usr/bin/env python3
"""
Mr.T 予想データ収集スクレイパー（netkeiba.com）

E0: Mr.T 2か月分データの DB・xlsx 蓄積

収集元: https://keirin.netkeiba.com/yoso/detail/?id=b{race_id}_634
- 634 = Mr.T（Mr.Tの競輪眼）の予想家ID
- race_id は連番ではなく飛び飛び（Mr.Tの予想があるIDのみ有効）

DoS対策:
  - rate_limit_sec=3.0 秒（デフォルト）
  - User-Agent: 正直な識別子
  - 1セッションあたりの最大リクエスト数制限あり

出力:
  - data/mrt_predictions.db (SQLite - マスター)
  - data/mrt_predictions.xlsx (エクスポート)

使い方:
  # 古い方から収集（12/28〜）
  python3 scripts/scrape_mrt_netkeiba.py --start-id 1479462 --end-id 1552200

  # 収集済みの続きから
  python3 scripts/scrape_mrt_netkeiba.py --resume

  # xlsxエクスポートのみ
  python3 scripts/scrape_mrt_netkeiba.py --export-only
"""

import argparse
import json
import os
import re
import sqlite3
import sys
import time
import urllib.error
import urllib.request
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# ─────────────────────────────────────────────────────────────
# 定数
# ─────────────────────────────────────────────────────────────

PREDICTOR_ID = 634  # Mr.T
BASE_URL = "https://keirin.netkeiba.com/yoso/detail/?id=b{race_id}_{predictor_id}"
RATE_LIMIT_SEC = 3.0
MAX_REQUESTS_PER_SESSION = 500
REQUEST_TIMEOUT = 15
USER_AGENT = "MrT-Collector/1.0 (keirin-prediction-research; rate-limited)"

# データベース
DB_DIR = Path(__file__).resolve().parent.parent / "data"
DB_PATH = DB_DIR / "mrt_predictions.db"
XLSX_PATH = DB_DIR / "mrt_predictions.xlsx"

# 連続404の打ち切り閾値
MAX_CONSECUTIVE_MISS = 5000


# ─────────────────────────────────────────────────────────────
# HTMLパーサー（軽量・外部依存なし）
# ─────────────────────────────────────────────────────────────

class NetkeibaHTMLParser(HTMLParser):
    """netkeiba.com の予想詳細ページからデータを抽出する。"""

    def __init__(self):
        super().__init__()
        self.text_chunks: List[str] = []
        self._current_tag = ""

    def handle_data(self, data):
        self.text_chunks.append(data.strip())

    def get_text(self) -> str:
        return "\n".join(c for c in self.text_chunks if c)


def fetch_page(race_id: int, predictor_id: int = PREDICTOR_ID) -> Optional[str]:
    """1ページ取得。存在しなければNone。"""
    url = BASE_URL.format(race_id=race_id, predictor_id=predictor_id)
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(req, timeout=REQUEST_TIMEOUT) as resp:
            return resp.read().decode("utf-8", errors="replace")
    except urllib.error.HTTPError as e:
        if e.code in (404, 400):
            return None
        raise
    except urllib.error.URLError:
        return None


def _strip_tags(html_fragment: str) -> str:
    """HTMLタグを除去してテキストのみ返す。"""
    return re.sub(r'<[^>]+>', ' ', html_fragment).strip()


def _extract_table_html(html: str, class_name: str) -> str:
    """指定クラス名のテーブルの生HTMLを返す。"""
    pattern = re.compile(
        rf'<table[^>]*{re.escape(class_name)}[^>]*>(.*?)</table>', re.S
    )
    m = pattern.search(html)
    return m.group(1) if m else ""


def _extract_table(html: str, class_name: str) -> str:
    """指定クラス名のテーブルHTMLを抽出し、タグを | 区切りのテキストに変換。"""
    raw = _extract_table_html(html, class_name)
    if not raw:
        return ""
    text = re.sub(r'<[^>]+>', '|', raw)
    text = re.sub(r'\|+', '|', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


# 予想印のCSSクラス → 記号マッピング
MARK_MAP = {
    "Icon_Honmei": "◎",
    "Icon_Taikou": "○",
    "Icon_Tanana": "▲",
    "Icon_Renka": "△",
    "Icon_Hoshi": "☆",
}


def parse_prediction_page(html: str, race_id: int) -> Optional[Dict[str, Any]]:
    """HTMLから予想データを構造化抽出する。

    netkeiba.com の主要HTMLクラス:
    - YosoShirushiTable01 内 <tr>: 予想印・選手
      - Icon_Honmei/Taikou/Tanana/Renka/Hoshi: 印
      - Waku{N}: 車番
      - PlayerName: 選手名
      - PlayerFrom: 県 期
      - PlayerClass: グレード 脚質
      - RaceCardCell01 内 数字: 競走得点
    - YosoKaimeTable01: 買い目テーブル
    - YosoRefundTable01: 払戻・収支
    - RaceResultTable 内 <tr>: 着順・車番・選手名
    - YosoKenkaiTitle: 見解タイトル（キャッチコピー）
    - YosoKenkaiTxt: 見解本文（展開予想全文）
    """
    if "指定した予想の取得が出来ませんでした" in html:
        return None
    if "getGoods Error" in html:
        return None

    data: Dict[str, Any] = {
        "race_id": race_id,
        "predictor_id": PREDICTOR_ID,
        "source_url": BASE_URL.format(race_id=race_id, predictor_id=PREDICTOR_ID),
    }

    # ─── 日付抽出（JS変数 var race_id = 202512284709） ───
    js_race_id = re.search(r"var\s+race_id\s*=\s*['\"]?(\d{12})['\"]?", html)
    if js_race_id:
        rid_str = js_race_id.group(1)
        data["date"] = f"{rid_str[:4]}-{rid_str[4:6]}-{rid_str[6:8]}"
    else:
        date_match = re.search(r'(\d{4})[年/.](\d{1,2})[月/.](\d{1,2})', html)
        if date_match:
            y, m, d = date_match.groups()
            data["date"] = f"{int(y)}-{int(m):02d}-{int(d):02d}"

    # ─── 会場名・レース番号（<title>から） ───
    title_match = re.search(
        r'<title[^>]*>\s*([\u4e00-\u9fff\u3040-\u309f\u30a0-\u30ff]+?)(\d{1,2})R',
        html
    )
    if title_match:
        data["venue"] = title_match.group(1)
        data["race_number"] = title_match.group(2) + "R"
    else:
        header = html[:len(html) // 5]
        venue_m = re.search(
            r'([\u4e00-\u9fff\u3040-\u309f\u30a0-\u30ff]{2,5})競輪', header
        )
        if venue_m and venue_m.group(1) not in ("ネット", "した", "の", "チケットミッドナイト"):
            data["venue"] = venue_m.group(1)
        race_m = re.search(r'(\d{1,2})R', header)
        if race_m:
            data["race_number"] = race_m.group(1) + "R"

    # ─── レースグレード ───
    upper = html[:len(html) // 3]
    for kw in ["決勝", "準決勝", "特選", "選抜", "予選", "一般"]:
        if kw in upper:
            data["race_grade"] = kw
            break
    for kw in ["S級", "A級", "L級"]:
        if kw in upper:
            data["race_class"] = kw
            break
    gm = re.search(r'(G[I]{1,3}|F[I12]{1,2})', upper)
    if gm:
        data["event_grade"] = gm.group(1)

    # ─── 選手データ（YosoShirushiTable01 の <tr> 単位） ───
    table_html = _extract_table_html(html, "YosoShirushiTable01")
    riders = []
    if table_html:
        rows = re.findall(r'<tr>(.*?)</tr>', table_html, re.S)
        for row in rows:
            # 予想印
            mark = ""
            for css_cls, symbol in MARK_MAP.items():
                if css_cls in row:
                    mark = symbol
                    break

            # 車番: <span class="Num Waku1">1</span>
            car_m = re.search(r'Waku(\d)', row)
            if not car_m:
                continue
            car_number = int(car_m.group(1))

            # 選手名: <dt class="PlayerName">丸山留依</dt>
            name_m = re.search(r'PlayerName">\s*(.*?)\s*<', row, re.S)
            name = name_m.group(1).strip() if name_m else ""

            # 県・期: <dd class="PlayerFrom">静岡 127期</dd>
            from_m = re.search(r'PlayerFrom">(.*?)<', row)
            prefecture, period = "", ""
            if from_m:
                pf = from_m.group(1).strip()
                pp = re.match(r'([\u4e00-\u9fff\u30a0-\u30ff]{1,4})\s+(\d{2,3})期', pf)
                if pp:
                    prefecture = pp.group(1)
                    period = pp.group(2) + "期"

            # グレード・脚質: <dd class="PlayerClass">Ａ２ 逃</dd>
            cls_m = re.search(r'PlayerClass">(.*?)<', row)
            grade, style = "", ""
            if cls_m:
                cs = cls_m.group(1).strip()
                # 全角→半角
                cs = cs.replace("Ａ", "A").replace("Ｓ", "S").replace("Ｌ", "L")
                cs = cs.replace("１", "1").replace("２", "2")
                gs = re.match(r'([AS][S12]?)\s+([逃追両])', cs)
                if gs:
                    grade = gs.group(1)
                    style = gs.group(2)

            # 得点: 93.55
            score_m = re.search(r'(\d{2,3}\.\d{2})', row)
            score = float(score_m.group(1)) if score_m else 0.0

            riders.append({
                "car_number": car_number,
                "name": name,
                "mark": mark,
                "prefecture": prefecture,
                "period": period,
                "grade": grade,
                "style": style,
                "score": score,
            })
    data["riders"] = riders

    # ─── 買い目（YosoKaimeTable01） ───
    kaime = _extract_table(html, "YosoKaimeTable01")
    bet_lines = []
    bet_type = ""
    if kaime:
        for bt in ["3連単", "3連複", "2車単", "2車複"]:
            if bt in kaime:
                bet_type = bt
                break
        # 3連系: |1|＞|4|＞|2| ... |1,200円|
        bet_pat = re.compile(
            r'\|(\d)\|[＞>]\|(\d)\|[＞>]\|(\d)\|.*?\|([\d,]+)円\|'
        )
        for m in bet_pat.finditer(kaime):
            bet_lines.append({
                "combination": f"{m.group(1)}→{m.group(2)}→{m.group(3)}",
                "amount": int(m.group(4).replace(",", "")),
            })
        # 2車系
        if not bet_lines and "2車" in bet_type:
            bet_pat2 = re.compile(
                r'\|(\d)\|[＞>]\|(\d)\|.*?\|([\d,]+)円\|'
            )
            for m in bet_pat2.finditer(kaime):
                bet_lines.append({
                    "combination": f"{m.group(1)}→{m.group(2)}",
                    "amount": int(m.group(3).replace(",", "")),
                })

    data["bet_type"] = bet_type or ("3連単" if "3連単" in html else "")
    data["bet_lines"] = bet_lines
    data["total_investment"] = sum(b["amount"] for b in bet_lines) if bet_lines else 10000

    # ─── 見解（YosoKenkaiTitle + YosoKenkaiTxt） ───
    # タイトル（キャッチコピー）
    title_m = re.search(r'<h2\s+class="YosoKenkaiTitle">(.*?)</h2>', html, re.S)
    kenkai_title = _strip_tags(title_m.group(1)).strip() if title_m else ""

    # 本文（展開予想全文）
    txt_m = re.search(r'<div\s+class="YosoKenkaiTxt">(.*?)</div>', html, re.S)
    kenkai_body = ""
    if txt_m:
        raw = txt_m.group(1)
        raw = re.sub(r'<br\s*/?\s*>', '\n', raw)  # 改行保持
        raw = re.sub(r'<[^>]+>', '', raw)
        kenkai_body = raw.strip()

    # タイトル + 本文を結合
    if kenkai_title and kenkai_body:
        data["comment"] = f"【{kenkai_title}】\n{kenkai_body}"
    elif kenkai_body:
        data["comment"] = kenkai_body
    elif kenkai_title:
        data["comment"] = kenkai_title
    else:
        # フォールバック
        for tb in re.findall(r'>([^<]{30,800})<', html):
            tb = tb.strip()
            if any(kw in tb for kw in ["展開", "主導権", "捲り", "番手", "カマシ", "狙い"]):
                data["comment"] = tb
                break
        else:
            data["comment"] = ""

    # ─── レース結果（RaceResultTable の <tr class="List"> 単位） ───
    result_html = _extract_table_html(html, "RaceResultTable")
    results = []
    if result_html:
        rows = re.findall(r'<tr\s+class="List">(.*?)</tr>', result_html, re.S)
        for row in rows:
            # 着順: <td class="ResultRank">1</td>
            pos_m = re.search(r'ResultRank">(\d+)<', row)
            # 車番: <td class="HorseNum Wakuban Waku1">1</td>
            car_m = re.search(r'Waku(\d)', row)
            # 選手名: <td class="HorseName">丸山留</td>
            name_m = re.search(r'HorseName">\s*(.*?)\s*<', row, re.S)
            if pos_m and car_m:
                results.append({
                    "position": int(pos_m.group(1)),
                    "car_number": int(car_m.group(1)),
                    "name": name_m.group(1).strip() if name_m else "",
                })
    data["results"] = results[:3]

    # ─── 払戻・収支（YosoRefundTable01） ───
    refund = _extract_table(html, "YosoRefundTable01")
    payout = 0
    if refund:
        payout_m = re.search(r'([\d,]+)円', refund)
        if payout_m:
            payout = int(payout_m.group(1).replace(",", ""))

    data["payout"] = payout
    data["profit_loss"] = payout - data["total_investment"]

    if "date" not in data:
        return None

    return data


# ─────────────────────────────────────────────────────────────
# DB操作
# ─────────────────────────────────────────────────────────────

def init_db(db_path: Path = DB_PATH) -> sqlite3.Connection:
    """DBを初期化し、テーブルを作成する。"""
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(db_path))
    conn.execute("""
        CREATE TABLE IF NOT EXISTS mrt_predictions (
            race_id INTEGER PRIMARY KEY,
            predictor_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            venue TEXT,
            race_number TEXT,
            race_grade TEXT,
            race_class TEXT,
            event_grade TEXT,
            bet_type TEXT,
            riders_json TEXT,
            bet_lines_json TEXT,
            total_investment INTEGER DEFAULT 10000,
            comment TEXT,
            results_json TEXT,
            payout INTEGER DEFAULT 0,
            profit_loss INTEGER DEFAULT 0,
            source_url TEXT,
            collected_at TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_mrt_date ON mrt_predictions(date)
    """)
    conn.commit()
    return conn


def save_prediction(conn: sqlite3.Connection, data: Dict[str, Any]) -> bool:
    """予想データをDBに保存。重複時はスキップ。"""
    try:
        conn.execute("""
            INSERT OR IGNORE INTO mrt_predictions (
                race_id, predictor_id, date, venue, race_number,
                race_grade, race_class, event_grade, bet_type,
                riders_json, bet_lines_json, total_investment,
                comment, results_json, payout, profit_loss,
                source_url, collected_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            data.get("race_id"),
            data.get("predictor_id", PREDICTOR_ID),
            data.get("date"),
            data.get("venue"),
            data.get("race_number"),
            data.get("race_grade"),
            data.get("race_class"),
            data.get("event_grade"),
            data.get("bet_type"),
            json.dumps(data.get("riders", []), ensure_ascii=False),
            json.dumps(data.get("bet_lines", []), ensure_ascii=False),
            data.get("total_investment", 10000),
            data.get("comment", ""),
            json.dumps(data.get("results", []), ensure_ascii=False),
            data.get("payout", 0),
            data.get("profit_loss", 0),
            data.get("source_url", ""),
            datetime.now().isoformat(),
        ))
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False  # 重複


def get_last_collected_id(conn: sqlite3.Connection) -> Optional[int]:
    """最後に収集したrace_idを返す。"""
    row = conn.execute("SELECT MAX(race_id) FROM mrt_predictions").fetchone()
    return row[0] if row and row[0] else None


def get_stats(conn: sqlite3.Connection) -> Dict[str, Any]:
    """収集統計を返す。"""
    row = conn.execute("""
        SELECT COUNT(*), MIN(date), MAX(date),
               SUM(CASE WHEN payout > 0 THEN 1 ELSE 0 END),
               SUM(payout), SUM(total_investment)
        FROM mrt_predictions
    """).fetchone()
    return {
        "total": row[0],
        "date_from": row[1],
        "date_to": row[2],
        "hits": row[3] or 0,
        "total_payout": row[4] or 0,
        "total_investment": row[5] or 0,
    }


# ─────────────────────────────────────────────────────────────
# xlsxエクスポート
# ─────────────────────────────────────────────────────────────

def export_xlsx(conn: sqlite3.Connection, xlsx_path: Path = XLSX_PATH):
    """DBからxlsxにエクスポート。openpyxlが必要。"""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl が必要です。 pip install openpyxl")
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mr.T予想"

    # ヘッダー
    headers = [
        "日付", "レース名", "グレード", "クラス", "券種",
        "予想印（選手・得点）", "買い目（組み合わせ・点数）",
        "投資額", "払戻金額", "収支", "レース結果", "見解"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # データ
    rows = conn.execute("""
        SELECT date, venue, race_number, race_grade, race_class,
               event_grade, bet_type, riders_json, bet_lines_json,
               total_investment, payout, profit_loss, results_json, comment
        FROM mrt_predictions
        ORDER BY date ASC, race_number ASC
    """).fetchall()

    for i, row in enumerate(rows, 2):
        date, venue, race_num, grade, rclass, egrade, btype, \
            riders_j, bets_j, invest, payout, pl, results_j, comment = row

        # レース名: 場名＋R
        race_name = f"{venue or ''}{race_num or ''}"

        # 予想印: 選手名(得点)
        riders = json.loads(riders_j) if riders_j else []
        riders_str = " / ".join(
            f"{r.get('car_number', '?')}番{r.get('name', '?')}({r.get('score', '?')})"
            for r in riders
        )

        # 買い目
        bets = json.loads(bets_j) if bets_j else []
        bets_str = " / ".join(
            f"{b.get('combination', '?')}={b.get('amount', '?')}円"
            for b in bets
        )

        # 結果
        results = json.loads(results_j) if results_j else []
        results_str = "→".join(
            f"{r.get('car_number', '?')}{r.get('name', '')}"
            for r in results
        )

        grade_str = " ".join(filter(None, [egrade, rclass, grade]))

        ws.cell(row=i, column=1, value=date)
        ws.cell(row=i, column=2, value=race_name)
        ws.cell(row=i, column=3, value=grade_str)
        ws.cell(row=i, column=4, value=rclass or "")
        ws.cell(row=i, column=5, value=btype or "3連単")
        ws.cell(row=i, column=6, value=riders_str)
        ws.cell(row=i, column=7, value=bets_str)
        ws.cell(row=i, column=8, value=invest)
        ws.cell(row=i, column=9, value=payout)
        ws.cell(row=i, column=10, value=pl)
        ws.cell(row=i, column=11, value=results_str)
        ws.cell(row=i, column=12, value=comment)

    # 列幅調整
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 50
    ws.column_dimensions["L"].width = 80

    wb.save(str(xlsx_path))
    print(f"xlsx exported: {xlsx_path} ({len(rows)} rows)")
    return True


# ─────────────────────────────────────────────────────────────
# メイン収集ループ
# ─────────────────────────────────────────────────────────────

def collect(start_id: int, end_id: int, conn: sqlite3.Connection,
            rate_limit: float = RATE_LIMIT_SEC,
            max_requests: int = MAX_REQUESTS_PER_SESSION) -> Dict[str, int]:
    """start_id → end_id の範囲でMr.T予想を収集する。

    クラスタジャンプ最適化:
    - Mr.T予想IDはクラスタ状に分布（同日予想はギャップ1-10で密集）
    - クラスタ間（日をまたぐ）は500-1500のギャップ
    - 50連続ミスでクラスタ終了と判断し、700IDスキップ
    - これにより全スキャン時間を約60%短縮
    """
    stats = {"fetched": 0, "saved": 0, "skipped": 0, "errors": 0,
             "miss": 0, "jumped": 0}
    consecutive_miss = 0
    in_cluster = True  # 常にジャンプ有効（初回から）
    CLUSTER_END_THRESHOLD = 20  # クラスタ終了判定（同日クラスタ内は最大ギャップ~10）
    JUMP_SIZE = 80  # クラスタ間ジャンプ量（最小ギャップ95に対応、見逃し防止）

    print(f"Collecting Mr.T predictions: b{start_id} → b{end_id}")
    print(f"Rate limit: {rate_limit}s, Max requests: {max_requests}")
    print(f"Cluster jump: {CLUSTER_END_THRESHOLD} miss → skip {JUMP_SIZE}")
    print("-" * 60)

    race_id = start_id
    while race_id <= end_id:
        if stats["fetched"] >= max_requests:
            print(f"\nMax requests ({max_requests}) reached. Stopping.")
            break

        # 進捗表示
        if stats["fetched"] % 50 == 0 and stats["fetched"] > 0:
            print(f"  Progress: {stats['fetched']} requests, "
                  f"{stats['saved']} saved, {stats['miss']} miss, "
                  f"jumps={stats['jumped']}, at=b{race_id}")

        # レート制限
        time.sleep(rate_limit)

        # ページ取得
        html = fetch_page(race_id)
        stats["fetched"] += 1

        if html is None:
            stats["miss"] += 1
            consecutive_miss += 1

            # クラスタ終了判定 → ジャンプ
            if in_cluster and consecutive_miss >= CLUSTER_END_THRESHOLD:
                jump_to = race_id + JUMP_SIZE
                if jump_to <= end_id:
                    print(f"  ⏭ Cluster end at b{race_id}, jumping to b{jump_to}")
                    race_id = jump_to
                    stats["jumped"] += 1
                    consecutive_miss = 0
                    continue

            if consecutive_miss >= MAX_CONSECUTIVE_MISS:
                print(f"\n{MAX_CONSECUTIVE_MISS} consecutive misses. Stopping.")
                break
            race_id += 1
            continue

        # パース
        prediction = parse_prediction_page(html, race_id)
        if prediction is None:
            stats["miss"] += 1
            consecutive_miss += 1
            if in_cluster and consecutive_miss >= CLUSTER_END_THRESHOLD:
                jump_to = race_id + JUMP_SIZE
                if jump_to <= end_id:
                    print(f"  ⏭ Cluster end at b{race_id}, jumping to b{jump_to}")
                    race_id = jump_to
                    stats["jumped"] += 1
                    consecutive_miss = 0
                    continue
            if consecutive_miss >= MAX_CONSECUTIVE_MISS:
                print(f"\n{MAX_CONSECUTIVE_MISS} consecutive misses. Stopping.")
                break
            race_id += 1
            continue

        # ヒット！
        consecutive_miss = 0
        in_cluster = True

        # DB保存
        saved = save_prediction(conn, prediction)
        if saved:
            stats["saved"] += 1
            venue = prediction.get("venue", "?")
            race_num = prediction.get("race_number", "?")
            date = prediction.get("date", "?")
            payout = prediction.get("payout", 0)
            print(f"  ✓ b{race_id}: {date} {venue}{race_num} "
                  f"payout=¥{payout:,}")
        else:
            stats["skipped"] += 1

        race_id += 1

    print("-" * 60)
    print(f"Done: {stats['fetched']} fetched, {stats['saved']} saved, "
          f"{stats['skipped']} duplicates, {stats['miss']} miss, "
          f"{stats['jumped']} jumps")
    return stats


# ─────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Mr.T 予想データ収集（netkeiba.com）"
    )
    parser.add_argument("--start-id", type=int, default=1479462,
                        help="開始race_id (default: 1479462 = 2025-12-28)")
    parser.add_argument("--end-id", type=int, default=1555000,
                        help="終了race_id")
    parser.add_argument("--resume", action="store_true",
                        help="前回の続きから収集")
    parser.add_argument("--export-only", action="store_true",
                        help="xlsxエクスポートのみ")
    parser.add_argument("--rate-limit", type=float, default=RATE_LIMIT_SEC,
                        help=f"リクエスト間隔(秒) (default: {RATE_LIMIT_SEC})")
    parser.add_argument("--max-requests", type=int, default=MAX_REQUESTS_PER_SESSION,
                        help=f"最大リクエスト数 (default: {MAX_REQUESTS_PER_SESSION})")
    parser.add_argument("--stats", action="store_true",
                        help="収集統計を表示")
    args = parser.parse_args()

    conn = init_db()

    if args.stats:
        s = get_stats(conn)
        print(f"Total: {s['total']} predictions")
        print(f"Date range: {s['date_from']} → {s['date_to']}")
        print(f"Hits: {s['hits']}/{s['total']} "
              f"({s['hits']/s['total']*100:.1f}% hit rate)" if s['total'] > 0 else "")
        print(f"Total payout: ¥{s['total_payout']:,}")
        print(f"Total investment: ¥{s['total_investment']:,}")
        if s['total_investment'] > 0:
            print(f"ROI: {s['total_payout']/s['total_investment']*100:.1f}%")
        conn.close()
        return

    if args.export_only:
        export_xlsx(conn)
        conn.close()
        return

    start_id = args.start_id
    if args.resume:
        last_id = get_last_collected_id(conn)
        if last_id:
            start_id = last_id + 1
            print(f"Resuming from b{start_id} (last collected: b{last_id})")
        else:
            print(f"No previous data. Starting from b{start_id}")

    stats = collect(start_id, args.end_id, conn,
                    rate_limit=args.rate_limit,
                    max_requests=args.max_requests)

    # 収集後にxlsxエクスポート
    if stats["saved"] > 0:
        export_xlsx(conn)

    # 最終統計
    s = get_stats(conn)
    print(f"\nDB total: {s['total']} predictions ({s['date_from']} → {s['date_to']})")

    conn.close()


if __name__ == "__main__":
    main()
